from flask import Flask,render_template,flash,redirect,url_for,session,logging,request,jsonify
from wtforms import Form,StringField,PasswordField,TextAreaField,validators
from flask_sqlalchemy import SQLAlchemy
import urllib
from functools import wraps
import pyodbc
from flask_table import Table,Col,LinkCol
from datetime import datetime
from openpyxl import *
from openpyxl.styles import Font
from win32com.client import Dispatch


app = Flask(__name__)

app.secret_key = "web_tool"

class MaintainerForm(Form):
    name = StringField("İsim")


class UpdateForm(Form):
    name = StringField("İsim")
    last_name = StringField("Soyisim")
    username = StringField("Kullanıcı Adı")
    password = StringField("Şifre")


class UpdateMaintenceForm(Form):
    date = StringField("Arızanın Meydana Geldiği Tarih:")
    last_date = StringField("Arızanın Meydana Giderildiği Tarih:")
    totalhour = StringField("Toplam Duruş:")
    section = StringField("Arızanın Meydana Geldiği Bölüm:")
    person = StringField("Arızayı Gideren Kişi 1:")
    person2 = StringField("Arızayı Gideren Kişi 2:")
    person3 = StringField("Arızayı Gideren Kişi 3:")
    person4 = StringField("Arızayı Gideren Kişi 4:")
    description = StringField("Arızanın Tanımı:")
    status = StringField("Durum:")
    operation = StringField("Yapılan İşlem:")


class RegisterForm(Form):
    name = StringField("İsim")
    last_name = StringField("Soyisim")
    username = StringField("Kullanıcı Adı")
    password = PasswordField("Şifre")
    status = StringField("Bu Kullanıcı Admin Mi?")

class OrderForm(Form):
    sp1 = StringField("Öğe 1")
    sp2 = StringField("Öğe 2")
    sp3 = StringField("Öğe 3")
    sp4 = StringField("Öğe 4")
    hm1 = StringField("Hammadde 1")
    hm2 = StringField("Hammadde 2")
    hm3 = StringField("Hammadde 3")
    hm4 = StringField("Hammadde 4")

class LoginForm(Form):
    username = StringField("Username")
    password = PasswordField("Password")

conn = pyodbc.connect("DRIVER={SQL Server};SERVER=VEEAMBACKUP\VEEAMSQL2016;PORT=1433;DATABASE=web_server;UID=sa;PWD=yasinc;")
conn2 = pyodbc.connect("DRIVER={SQL Server};SERVER=VEEAMBACKUP\VEEAMSQL2016;PORT=1433;DATABASE=web_server;UID=sa;PWD=yasinc;")
cursor2 = conn2.cursor()
cursor = conn.cursor()

@app.route("/")
def welcomepage():
    if "logged_in" in session:
        return render_template("index.html")
    else:
        return render_template("welcomepage.html")

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("welcomepage"))

@app.route("/login", methods = ["GET", "POST"])
def login():
    form = LoginForm(request.form)
    if request.method == "POST":
        u_name = form.username.data
        password_entered = form.password.data

        cursor = conn.cursor()

        query = "SELECT password FROM users WHERE username=?"
        cursor.execute(query,[u_name])
        password = cursor.fetchall()
        for row in password:
            if row[0] == password_entered:
                cursor = conn.cursor()
                commandstring = "SELECT name,last_name,is_admin FROM users WHERE username=?"
                cursor.execute(commandstring,[u_name])
                data = cursor.fetchall()
                for elements in data:
                    session["name"] = elements[0]
                    session["last_name"] = elements[1]
                    session["logged_in"] = True
                    session["is_admin"] = elements[2]
                    return redirect(url_for("index"))
    else:
        return render_template("login.html", form = form)

def login_required(f):
    @wraps(f)
    def decorated_function(*args,**kwargs):
        if "logged_in" in session:
            return f(*args, **kwargs)
        else:
            flash("Bu Araçlara Ulaşabilmek İçin Giriş Yapmalısınız!","danger")
            return redirect(url_for('login'))
    return decorated_function

@app.route("/index")
@login_required
def index():
    return render_template("index.html")

@app.route("/calculations")
@login_required
def calculations():
    return render_template("calculations.html")

@app.route("/deleteuser/<string:username>")
@login_required
def user(username):
    if session["is_admin"] == "yes":
        cursor = conn.cursor()
        query = "DELETE FROM users WHERE username=?"
        cursor.execute(query,[username])
        cursor.commit()
        return redirect(url_for("useradministration"))
    else:
        flash("Bu sayfaya Sadece yetkili kullanıcılar girebilir. Lütfen denemeyiniz!","danger")
        return redirect(url_for("index"))

@app.route("/useradministration")
@login_required
def useradministration():
    if session["is_admin"] == "yes":
        cursor.execute("select id,name,last_name,username,password from users")
        data = cursor.fetchall() 
        return render_template("useradministration.html",value=data)      
    else:
        flash("Bu sayfaya Sadece yetkili kullanıcılar girebilir. Lütfen denemeyiniz!","danger")
        return redirect(url_for("index"))
        

@app.route("/adduser", methods = ["GET", "POST"])
@login_required
def adduser():
    form = RegisterForm(request.form)
    if session["is_admin"] == "yes":
        if request.method == "GET":
            return render_template("adduser.html", form = form)
        else:
            cursor = conn.cursor()
            query = "INSERT INTO users (name,last_name,username,password,is_admin) VALUES (?,?,?,?,?)"
            name = form.name.data
            last_name = form.last_name.data
            username = form.username.data
            password = form.password.data
            is_admin = form.status.data
            cursor.execute(query,[name,last_name,username,password,is_admin])
            cursor.commit()
            return redirect(url_for("useradministration"))
    else:
        return redirect(url_for("index"))

@app.route("/updateuser/<string:id>", methods = ["GET", "POST"])
@login_required
def updateuser(id):
    if session["is_admin"] == "yes":
        form = UpdateForm(request.form)
        if request.method == "GET":
            cursor = conn.cursor()
            query = "SELECT * FROM users WHERE id =?"
            cursor.execute(query,[id])
            info = cursor.fetchone()
            form.name.data = info[1]
            form.last_name.data = info[2]
            form.username.data = info[3]
            form.password.data = info[4]
            return render_template("updateuser.html", form = form)
        else:
            name = form.name.data
            last_name = form.last_name.data
            username = form.username.data
            password = form.password.data
            cursor = conn.cursor()
            query = "UPDATE users SET name=?,last_name=?,username=?,password=? WHERE id=?"
            cursor.execute(query,[name,last_name,username,password,id])
            return redirect(url_for("useradministration"))
    else:
        return redirect(url_for("index"))
@app.route("/otomaticorder")
@login_required
def order():
    form = OrderForm(request.form)
    return render_template("otomaticorder.html", form = form)

@app.route("/maintenance")
@login_required
def maintenancemain():
        total = "" 
        cursor.execute("SELECT * FROM maintenance")
        data = cursor.fetchall()
        cursor2.execute("SELECT name,maintenanceid FROM maintainers")
        data2 = cursor2.fetchall()
        while data2[1] == data[0]:
            total = total + " " + data2[0]
        return render_template("maintenance/maintenancemain.html",value=data, value2=data2, bruh=total)

@app.route("/addmaintenanceform", methods = ["GET", "POST"])
@login_required
def maintenanceadd():
    form = UpdateMaintenceForm(request.form)
    if request.method == "GET":
        return render_template("maintenance/addmaintenance.html",form = form)  
    else:
        datestring = str(form.date.data)
        last_datestring = str(form.last_date.data)
        date  = datetime.strptime(datestring,'%d-%m-%Y')
        last_date  = datetime.strptime(last_datestring,'%d-%m-%Y')
        totalhour = form.totalhour.data
        section = form.section.data
        person = form.person.data
        description = form.description.data
        status = form.status.data
        operation = form.operation.data
        cursor = conn.cursor()
        query = "INSERT INTO maintenance (date,last_date,totalhour,section,person,description,status,operation) VALUES (?,?,?,?,?,?,?,?)"
        cursor.execute(query,[date,last_date,totalhour,section,person,description,status,operation])
        cursor.commit()
        return redirect(url_for("maintenancemain"))

@app.route("/deletemaintenanceform/<string:id>")
@login_required
def deletemaintenanceform(id):  
    cursor = conn.cursor()
    query = "DELETE FROM maintenance WHERE id=?"
    cursor.execute(query,[id])
    cursor.commit()
    return redirect(url_for("maintenancemain"))


@app.route("/updatemaintenanceform/<string:id>", methods = ["GET","POST"])
@login_required
def updatemaintenanceform(id):
    form = UpdateMaintenceForm(request.form)
    if request.method == "GET":
        cursor = conn.cursor()
        query = "SELECT * FROM maintenance WHERE id =?"
        cursor.execute(query,[id])
        info = cursor.fetchone()
        form.date.data = info[1]
        form.last_date.data = info[2]
        form.totalhour.data = info[3]
        form.section.data = info[4]
        form.person.data = info[5]
        form.person2.data = info[9]
        form.person3.data = info[10]
        form.person4.data = info[11]
        form.description.data = info[6]
        form.status.data = info[7]
        form.operation.data = info[8]
        return render_template("maintenance/maintenanceupdate.html", form = form)
    else:
        date = form.date.data
        last_date = form.last_date.data
        totalhour = form.totalhour.data
        section = form.section.data
        person1 = form.person.data
        person2 = form.person2.data
        person3 = form.person3.data
        person4 = form.person4.data
        description = form.description.data
        status = form.status.data
        operation = form.operation.data
        cursor = conn.cursor()
        query = "UPDATE maintenance SET date=?,last_date=?,totalhour=?,section=?,person=?,description=?,status=?,operation=?,person_2=?,person_3=?,person_4=? WHERE id=?"
        cursor.execute(query,[date,last_date,totalhour,section,person1,description,status,operation,person2,person3,person4,id])
        cursor.commit()
        return redirect(url_for("maintenancemain"))

@app.route("/excelmaintenance/<string:id>")
@login_required
def excelmaintenance(id):

    row = 1
    column = 2
    workbook = Workbook()
    sheet = workbook.active
    bold_font = Font(bold=True, size=20)
    value_font = Font(size=15)


    sheet.cell(row=row, column=column).value = "Sıra No"
    sheet.cell(row=row, column=column + 1).value = "Arızanın Meydana Geldiği Tarih"
    sheet.cell(row=row, column=column + 2).value = "Arızanın Giderildiği Tarih"
    sheet.cell(row=row, column=column + 3).value = "Toplam Duruş (saat)"
    sheet.cell(row=row, column=column + 4).value = "Arızanın Meydana Geldiği Bölüm"
    sheet.cell(row=row, column=column + 5).value = "Arızayı Gideren Kişi 1"
    sheet.cell(row=row, column=column + 6).value = "Arızayı Gideren Kişi 2"
    sheet.cell(row=row, column=column + 7).value = "Arızayı Gideren Kişi 3"
    sheet.cell(row=row, column=column + 8).value = "Arızayı Gideren Kişi 4"
    sheet.cell(row=row, column=column + 9).value = "Arıza Tanımı"
    sheet.cell(row=row, column=column + 10).value = "Durum"
    sheet.cell(row=row, column=column + 11).value = "Yapılan İşlem"

    sheet.cell(row=row, column=column).font = bold_font
    sheet.cell(row=row, column=column + 1).font = bold_font
    sheet.cell(row=row, column=column + 2).font = bold_font
    sheet.cell(row=row, column=column + 3).font = bold_font
    sheet.cell(row=row, column=column + 4).font = bold_font
    sheet.cell(row=row, column=column + 5).font = bold_font
    sheet.cell(row=row, column=column + 6).font = bold_font
    sheet.cell(row=row, column=column + 7).font = bold_font
    sheet.cell(row=row, column=column + 8).font = bold_font
    sheet.cell(row=row, column=column + 9).font = bold_font
    sheet.cell(row=row, column=column + 10).font = bold_font
    sheet.cell(row=row, column=column + 11).font = bold_font
    
    cursor = conn.cursor()
    query = "SELECT * FROM maintenance WHERE id=?"
    cursor.execute(query,[id])

    data = cursor.fetchone()

    sheet.cell(row=row + 1, column=column).value = data[0]
    sheet.cell(row=row + 1, column=column + 1).value = str(data[1])
    sheet.cell(row=row + 1, column=column + 2).value = str(data[2])
    sheet.cell(row=row + 1, column=column + 3).value = data[3]
    sheet.cell(row=row + 1, column=column + 4).value = data[4]
    sheet.cell(row=row + 1, column=column + 5).value = data[5]
    sheet.cell(row=row + 1, column=column + 6).value = data[9]
    sheet.cell(row=row + 1, column=column + 7).value = data[10]
    sheet.cell(row=row + 1, column=column + 8).value = data[11]
    sheet.cell(row=row + 1, column=column + 9).value = data[6]
    sheet.cell(row=row + 1, column=column + 10).value = data[7]
    sheet.cell(row=row + 1, column=column + 11).value = data[8]

    sheet.cell(row=row + 1, column=column).font = value_font
    sheet.cell(row=row + 1, column=column + 1).font = value_font
    sheet.cell(row=row + 1, column=column + 2).font = value_font
    sheet.cell(row=row + 1, column=column + 3).font = value_font
    sheet.cell(row=row + 1, column=column + 4).font = value_font
    sheet.cell(row=row + 1, column=column + 5).font = value_font
    sheet.cell(row=row + 1, column=column + 6).font = value_font
    sheet.cell(row=row + 1, column=column + 7).font = value_font
    sheet.cell(row=row + 1, column=column + 8).font = value_font
    sheet.cell(row=row + 1, column=column + 9).font = value_font
    sheet.cell(row=row + 1, column=column + 10).font = value_font
    sheet.cell(row=row + 1, column=column + 11).font = value_font

    filename = "Bakım Formu Numara " + str(data[0]) + ".xlsx"

    workbook.save(filename=filename)

    workbook = load_workbook(filename = filename)

    xl = Dispatch("Excel.Application")
    xl.Visible = True
    wb = xl.Workbooks.Open(r'C:\Users\owner\Desktop\web_tool\Bakım Formu Numara ' + str(data[0]) + '.xlsx')
    return redirect(url_for("maintenancemain"))
        
    



if __name__ == "__main__":
    app.run(port=7000, host=("192.168.10.44"),debug=True)